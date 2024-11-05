import shutil
import requests
from PIL import Image as PilImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import os

# Function to load the Excel template and prepare for modification
def load_template(template_path, new_file_path):
    # Copy the template to a new file
    shutil.copy(template_path, new_file_path)
    # Load the new workbook
    wb = load_workbook(new_file_path)
    ws = wb.active
    return wb, ws

# Function to insert series data starting from row 4
def insert_series_data(ws, series_data, start_row=4):
    for i, item in enumerate(series_data):
        row = start_row + i
        print(f"Inserting series data into row {row}")
        print(item)
        ws[f"C{row}"] = item.get("物品名称", "")
        ws[f"D{row}"] = item.get("单位", "")
        ws[f"E{row}"] = item.get("数量", "")
        ws[f"F{row}"] = item.get("单价", "")
        ws[f"G{row}"] = item.get("用途说明", "")
        ws[f"H{row}"] = item.get("备注", "")
        if row >= 8:  # Stop if the row exceeds row 9
            break

# Function to save the workbook after modification
def save_workbook(wb, new_file_path):
    wb.save(new_file_path)

# Main function to generate the Excel file
def generate_excel_file(template_path, new_file_path, order_data):
    print(f"Generating Excel file for order {order_data.get('订单信息', '')}")
    # Load template
    wb, ws = load_template(template_path, new_file_path)
    
    # Insert order details into specific cells
    ws["B6"] = order_data.get("customer_id", "")
    
    # # Insert series data from row 4 to 9
    # insert_series_data(ws, order_data.get("seriesData", []))

    # Save the workbook
    save_workbook(wb, new_file_path)
    
    print(f"Workbook saved as {new_file_path}")

# # Example usage
# if __name__ == "__main__":
#     # Example parameters
#     template_path = "./production_order_sheet.xlsx"  # Replace with your template path
#     new_file_path = "./production_order_sheet_modified.xlsx"  # The output file
#     image_url = "http://localhost:12667/shoe/3E3056/shoe_image.jpg"
#     image_save_path = "shoe_image.jpg"  # The image will be downloaded to this path

#     # Example order details
#     order_details = {
#         'order_id': '12345',
#         'input_person': 'John Doe',
#         'order_finish_time': '2024-09-01',
#         'inherit_id': '98765',
#         'customer_id': 'CUST123',
#         'last_type': 'TypeX'
#     }

#     # Call the function to generate the Excel file
#     generate_excel_file(template_path, new_file_path, order_details)
