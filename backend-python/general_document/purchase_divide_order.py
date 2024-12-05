import shutil
from openpyxl import load_workbook
import os
from openpyxl.styles import Border, Side

# Function to load the Excel template and prepare for modification
def load_template(template_path, new_file_path):
    # Copy the template to a new file
    shutil.copy(template_path, new_file_path)
    # Load the new workbook
    wb = load_workbook(new_file_path)
    ws = wb.active
    return wb, ws
def add_borders(ws, start_cell, end_cell):
    thin = Side(border_style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Iterate over the range and apply borders
    for row in ws[start_cell:end_cell]:
        for cell in row:
            cell.border = border
# Function to insert series data starting from row 4
def merge_cells(ws, row):
    ws.merge_cells(f"A{row+1}:D{row+1}")
    ws.merge_cells(f"B{row+2}:H{row+2}")
    ws.merge_cells(f"B{row+3}:H{row+3}")
    ws.merge_cells(f"B{row+4}:D{row+4}")
    ws.merge_cells(f"E{row+4}:H{row+4}")
    ws.merge_cells(f"A3:A{row}")
    add_borders(ws, f"A3", f"H{row+1}")
def insert_series_data(ws, series_data, start_row=4):
    required_rows = 5  # Minimum rows to keep
    row = start_row - 1  # To calculate the last row after loop
    for i, item in enumerate(series_data):
        row = start_row + i
        print(f"Inserting series data into row {row}")
        print(item)
        ws[f"B{row}"] = i + 1
        ws[f"C{row}"] = item.get("物品名称", "")
        ws[f"D{row}"] = item.get("单位", "")
        ws[f"E{row}"] = item.get("数量", "")
        ws[f"F{row}"] = item.get("单价", "")
        ws[f"G{row}"] = item.get("用途说明", "")
        ws[f"H{row}"] = item.get("备注", "")
    for i in range(len(series_data), required_rows):
        row = start_row + i
        print(f"Adding empty row at {row}")
        ws[f"B{row}"] = i + 1  # Continue numbering for empty rows

    return row

# Function to save the workbook after modification
def save_workbook(wb, new_file_path):
    wb.save(new_file_path)

# Main function to generate the Excel file
def generate_excel_file(template_path, new_file_path, order_data):
    print(f"Generating Excel file for order {order_data.get('订单信息', '')}")
    # Load template
    wb, ws = load_template(template_path, new_file_path)
    
    # Insert order details into specific cells
    ws["D2"] = order_data.get("订单信息", "")
    ws["B2"] = order_data.get("供应商", "")
    ws["H2"] = order_data.get("日期", "")
    
    # Insert series data from row 4 to 9
    row = insert_series_data(ws, order_data.get("seriesData", []))
    ws[f"A{row+1}"] = "合计"
    ws[f"E{row+1}"] = order_data.get("合计", "")
    ws[f"F{row+1}"] = order_data.get("备注", "")
    ws[f"A{row+2}"] = "环境要求:"
    ws[f"A{row+3}"] = "发货地址:"
    ws[f"A{row+4}"] = "交货期限:"
    ws[f"B{row+2}"] = order_data.get("环保要求", "")
    ws[f"B{row+3}"] = order_data.get("发货地址", "")
    ws[f"B{row+4}"] = order_data.get("交货期限", "")
    ws[f"E{row+4}"] = "如有特殊情况提前5天反馈，无故延期有贵公司承担后续责任。"
    ws[f"A{row+5}"] = "制表:"
    ws[f"E{row+5}"] = "审核:"

    merge_cells(ws, row)

    # Save the workbook
    save_workbook(wb, new_file_path)
    
    print(f"Workbook saved as {new_file_path}")

# def test_case_1():
#     template_path = "D:/catSupermarket/jiancheng/backend-python/general_document/标准采购订单.xlsx"
#     new_file_path = "D:/catSupermarket/jiancheng/backend-python/general_document/pur_test.xlsx"
#     order_data = {
#         "订单信息": "订单编号12345",
#         "供应商": "供应商A",
#         "日期": "2024-11-19",
#         "seriesData": [
#             {"物品名称": "商品1", "单位": "个", "数量": 10, "单价": 5.5, "用途说明": "用途1", "备注": "备注1"},
#             {"物品名称": "商品2", "单位": "箱", "数量": 20, "单价": 15.0, "用途说明": "用途2", "备注": "备注2"},
#         ],
#         "合计": 350,
#         "备注": "总备注",
#         "环保要求": "符合环保标准",
#         "发货地址": "测试地址",
#         "交货期限": "2024-12-01",
#     }
#     generate_excel_file(template_path, new_file_path, order_data)
#     print("Test Case 1: Basic functionality passed.")

