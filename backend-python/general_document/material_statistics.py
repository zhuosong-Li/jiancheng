from openpyxl.utils import range_boundaries

from openpyxl import load_workbook
def generate_material_statistics_file(template_path, save_path, order_rid, order_shoe_rid, customer_name, materials_data):
    """
    Generates a 材料统计表 file based on the provided template and materials data.
    
    Parameters:
    - template_path: Path to the template file.
    - save_path: Path to save the generated file.
    - order_rid: Order RID to place in cell B2.
    - order_shoe_rid: Order Shoe RID to place in cell D2.
    - customer_name: Customer name to place in cell F2.
    - materials_data: List of dictionaries containing supplier_name, material_name, model, specification, approval_amount, and purchase_amount.
    """
    # Load the workbook and select the first sheet
    workbook = load_workbook(template_path)
    sheet = workbook["Sheet1"]
    
    # Insert metadata
    sheet["B2"] = order_rid
    sheet["D2"] = order_shoe_rid
    sheet["F2"] = customer_name
    
    # Manage merged cell in H2
    for merged_cell_range in sheet.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(merged_cell_range))
        if min_row == 2 and min_col == 8:  # Check if it's H2
            sheet.unmerge_cells(str(merged_cell_range))
            sheet["H2"] = datetime.now().strftime("%Y-%m-%d")  # Insert date
            sheet.merge_cells(start_row=min_row, start_column=min_col, end_row=max_row, end_column=max_col)
            break

    # Insert materials data starting from row 3
    start_row = 4
    for index, data in enumerate(materials_data, start=start_row):
        supplier_name = data.get("supplier_name", "")
        material_name = data.get("material_name", "")
        model = data.get("model", "")
        specification = data.get("specification", "")
        approval_amount = data.get("approval_amount", 0)
        purchase_amount = data.get("purchase_amount", 0)
        print(supplier_name, material_name, model, specification, approval_amount, purchase_amount)
        print(index)
        
        # Fill the cells
        sheet[f"A{index}"] = supplier_name
        sheet[f"B{index}"] = f"{material_name} {model} {specification}".strip()
        sheet[f"C{index}"] = approval_amount
        sheet[f"E{index}"] = purchase_amount

    # Save the modified file
    workbook.save(save_path)
    return save_path

