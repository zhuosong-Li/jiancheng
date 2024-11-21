import shutil
import requests
from PIL import Image as PilImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import os

# Function to load the Excel template and prepare for modification
def load_template(template_path, new_file_path):
    # Copy the template to a new file
    shutil.copy(template_path, new_file_path)
    # Load the new workbook
    wb = load_workbook(new_file_path)
    ws = wb.active
    return wb, ws

# Function to download an image from a URL and save it locally
def download_image(image_path, save_path):
    # Check if the source image exists
    if os.path.exists(image_path):
        # Copy the image to the destination
        shutil.copy(image_path, save_path)
        print(f"Image copied successfully from {image_path} to {save_path}")
    else:
        print(f"Image not found at: {image_path}")

# Function to calculate the size of the image to fit within a cell range
def resize_image_to_fit(ws, image_path, start_cell, end_cell):
    # Get the width and height of the target cell range
    col_start, row_start = ws[start_cell].column, ws[start_cell].row
    col_end, row_end = ws[end_cell].column, ws[end_cell].row

    # Calculate width and height in pixels
    width_in_pixels = sum([ws.column_dimensions[get_column_letter(col)].width or 8.43 for col in range(col_start, col_end + 1)]) * 7
    height_in_pixels = sum([ws.row_dimensions[row].height or 15 for row in range(row_start, row_end + 1)])

    # Open the image
    img = PilImage.open(image_path)

    # Convert the image to RGB if it has an alpha channel (RGBA)
    if img.mode == 'RGBA':
        img = img.convert('RGB')

    # Resize the image to fit within the calculated width and height
    img = img.resize((int(width_in_pixels), int(height_in_pixels)), PilImage.Resampling.LANCZOS)
    
    # Create the resized image path
    directory, filename = os.path.split(image_path)
    name, ext = os.path.splitext(filename)
    resized_image_path = os.path.join(directory, f"resized_{name}{ext}")
    
    # Save the resized image
    img.save(resized_image_path)

    return resized_image_path

def insert_order_details(ws, order_details):
    ws['C4'] = order_details.get('order_id')
    ws['I4'] = order_details.get('inherit_id')
    ws['L4'] = order_details.get('customer_id')
    ws['I5'] = order_details.get('last_type')
    ws['L5'] = order_details.get('origin_size')
    ws['O5'] = order_details.get('size_range')
    ws['I6'] = order_details.get('size_difference')
    ws['L6'] = order_details.get('designer')
    ws['O6'] = order_details.get('brand')
    ws['I7'] = order_details.get('colors')
def insert_series_data(ws, series_data, start_row=11):
    for i, row_data in enumerate(series_data):
        for col_index, (col, value) in enumerate(row_data.items()):
            column_map = {
                "鞋型颜色": "A", "材料类型": "B","材料二级类型": "C", "材料名称":"E", "材料型号": "G", "材料规格": "I",
                "颜色": "K", "单位": "L", "厂家名称": "M", "备注": "O"
            }
            column_letter = column_map.get(col)
            if column_letter:
                ws[f"{column_letter}{start_row + i}"] = value
def insert_image(ws, image_path, start_cell, end_cell):
    # Resize the image to fit the target cell range
    resized_image_path = resize_image_to_fit(ws, image_path, start_cell, end_cell)
    
    # Insert the resized image into the worksheet
    img = Image(resized_image_path)
    ws.add_image(img, start_cell)
def save_workbook(wb, new_file_path):
    wb.save(new_file_path)
def generate_instruction_excel_file(template_path, new_file_path, order_details, series_data, image_url, image_save_path):
    # Load template
    wb, ws = load_template(template_path, new_file_path)

    # Insert the order details
    insert_order_details(ws, order_details)

    # Insert series data starting from row 9
    insert_series_data(ws, series_data, start_row=11)
    # Download the image from the URL
    download_image(image_url, image_save_path)
    if os.path.exists(image_save_path):
        insert_image(ws, image_save_path, "A6", "F8")

    # Save the modified workbook
    save_workbook(wb, new_file_path)

    print(f"Workbook saved as {new_file_path}")
