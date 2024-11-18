import shutil
from openpyxl import load_workbook


# Function to load the Excel template and prepare for modification
def load_template(template_path, new_file_path):
    # Copy the template to a new file
    shutil.copy(template_path, new_file_path)
    # Load the new workbook
    wb = load_workbook(new_file_path)
    ws = wb.active
    return wb, ws


# Function to insert series data starting from row 4
def insert_series_data(ws, series_data, start_row=4):
    for i, item in enumerate(series_data):
        row = start_row + (i * 2)
        ws[f"A{row}"] = item.get("row_id", "")
        ws[f"B{row}"] = item.get("procedure_name", "")


# Function to save the workbook after modification
def save_workbook(wb, new_file_path):
    wb.save(new_file_path)


# Main function to generate the Excel file
def generate_excel_file(template_path, new_file_path, data: dict):
    print(f"Generating Excel file")
    print(data)
    # Load template
    wb, ws = load_template(template_path, new_file_path)

    # create new sheet for every 20 rows
    table_data = data.get("procedures", [])
    table_data = table_data * 7
    for i in range(0, len(table_data), 20):
        if i > 0:
            copied_ws = wb.copy_worksheet(ws)
            copied_ws.title = f"Sheet{(i//20)+1}"

    for i in range(0, len(table_data), 20):
        new_list = table_data[i : i + 20]
        # current worksheet
        ws = wb.worksheets[i // 20]
        # Insert order details into specific cells
        ws["D2"] = data.get("shoe_rid", "")
        insert_series_data(ws, new_list)

    # Save the workbook
    save_workbook(wb, new_file_path)
    print(f"Workbook saved as {new_file_path}")
