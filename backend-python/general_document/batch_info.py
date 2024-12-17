import shutil
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string

COLUMNS = 0
# Function to load the Excel template and prepare for modification
def load_template(template_path, new_file_path):
    # Copy the template to a new file
    shutil.copy(template_path, new_file_path)
    # Load the new workbook
    wb = load_workbook(new_file_path)
    return wb


def get_next_column_name(current_column_name):
    # Convert column letter to index
    column_index = column_index_from_string(current_column_name)
    # Increment the index to get the next column
    next_column_index = column_index + 1
    # Convert the new index back to a column letter
    next_column_name = get_column_letter(next_column_index)
    return next_column_name


# Function to copy formatting and insert a new row
def insert_row_with_format(ws, row_to_copy, new_row_idx):
    # Insert a new row
    ws.insert_rows(new_row_idx)
    # Copy the formatting
    for col_idx, cell in enumerate(ws[row_to_copy], start=1):
        new_cell = ws.cell(row=new_row_idx, column=col_idx)
        if cell.has_style:
            new_cell.border = cell.border.copy()
            new_cell.alignment = cell.alignment.copy()


# Function to insert series data starting from row 6
def insert_series_data(wb: Workbook, series_data, column_amount_mapping, start_row=6):
    global COLUMNS
    ws = wb.active
    # un-merge cell
    ws.unmerge_cells("A7:F7")
    ws.unmerge_cells("A8:F8")
    for i, val in enumerate(series_data):
        new_row = i + start_row
        if i > 0:
            insert_row_with_format(ws, start_row, new_row)
        ws[f"D{new_row}"] = val["color_name"]
        ws[f"E{new_row}"] = val["packaging_info_name"]
        column_name = "F"
        for j in range(34, 47):
            if column_amount_mapping[j]:
                ws[f"{column_name}{new_row}"] = val[f"size_{j}_ratio"]
                cell = ws["F6"]
                if i == 0:
                    source_width = ws.column_dimensions["F"].width
                    ws.column_dimensions[column_name].width = source_width
                    column_index = column_index_from_string(column_name)
                    new_cell = ws.cell(row=start_row, column=column_index)
                    if cell.has_style:
                        new_cell.border = cell.border.copy()
                        new_cell.alignment = cell.alignment.copy()
                column_name = get_next_column_name(column_name)
        val["packaging_total_quantity"] = val["total_quantity_ratio"] * val["packaging_info_quantity"]
        for name in ["total_quantity_ratio", "packaging_info_quantity", "packaging_total_quantity"]:
            # insert total_quantity_ratio and packaging_info_quantity
            ws[f"{column_name}{new_row}"] = val[name]
            column_index = column_index_from_string(column_name)
            new_cell = ws.cell(row=start_row, column=column_index)
            if cell.has_style:
                new_cell.border = cell.border.copy()
                new_cell.alignment = cell.alignment.copy()
            column_name = get_next_column_name(column_name)
    # merge cells
    target_column = get_column_letter(COLUMNS)
    ws.merge_cells(f"A{6+len(series_data)}:{target_column}{6+len(series_data)}")
    ws.merge_cells(f"A{6+len(series_data) + 1}:{target_column}{6+len(series_data) + 1}")


# Function to save the workbook after modification
def save_workbook(wb, new_file_path):
    wb.save(new_file_path)


# Main function to generate the Excel file
def generate_excel_file(template_path, new_file_path, data: dict):
    global COLUMNS
    print(f"Generating Excel file")
    # Load template
    wb = load_template(template_path, new_file_path)

    # calculate how many columns
    length = 0
    for _, value in data["column_amount_mapping"].items():
        if value:
            length += 1
    COLUMNS = 5 + length + 3

    table_data = data["batch_info"]
    insert_series_data(wb, table_data, data["column_amount_mapping"])

    ws = wb.active
    # insert meta data
    brand = "商标：" + data["customer_brand"]
    ws["E4"] = brand
    ws.column_dimensions["E"].width = len(brand) + 10

    ws.merge_cells(f"A6:A{6+len(table_data)-1}")
    ws["A6"] = data["order_rid"]

    ws.merge_cells(f"B6:B{6+len(table_data)-1}")
    ws["B6"] = data["shoe_rid"]

    ws.merge_cells(f"C6:C{6+len(table_data)-1}")
    ws["C6"] = data["customer_product_name"]

    column_name = "F"
    f5_cell = ws["F5"]
    for i in range(0, len(data["shoe_size_names"])):
        if data["column_amount_mapping"][i+34]:
            ws[f"{column_name}{5}"] = data["shoe_size_names"][i]["label"]
            column_index = column_index_from_string(column_name)
            new_cell = ws.cell(row=5, column=column_index)
            if f5_cell.has_style:
                new_cell.border = f5_cell.border.copy()
                new_cell.alignment = f5_cell.alignment.copy()
            column_name = get_next_column_name(column_name)

    for name in ["对/件", "件数", "双数"]:
        ws[f"{column_name}{5}"] = name
        column_index = column_index_from_string(column_name)
        new_cell = ws.cell(row=5, column=column_index)
        if f5_cell.has_style:
            new_cell.border = f5_cell.border.copy()
            new_cell.alignment = f5_cell.alignment.copy()
        column_name = get_next_column_name(column_name)

    target_column = get_column_letter(COLUMNS)
    ws.merge_cells(f"A1:{target_column}1")
    ws.merge_cells(f"A2:{target_column}2")

    target_column = get_column_letter(COLUMNS)
    ws.merge_cells(f"F4:{target_column}4")

    # Save the workbook
    save_workbook(wb, new_file_path)
    print(f"Workbook saved as {new_file_path}")
