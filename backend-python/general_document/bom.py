import shutil
import requests
from PIL import Image as PilImage
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import os

# Function to load the Excel template and prepare for modification
def load_template(template_path, new_file_path):
    # Copy the template to a new file
    shutil.copy(template_path, new_file_path)
    # Load the new workbook
    wb = load_workbook(new_file_path)
    ws = wb.active
    return wb, ws

# Function to download an image from a URL and save it locally
def download_image(image_path, save_path):
    # Check if the source image exists
    if os.path.exists(image_path):
        # Copy the image to the destination
        shutil.copy(image_path, save_path)
        print(f"Image copied successfully from {image_path} to {save_path}")
    else:
        print(f"Image not found at: {image_path}")

# Function to calculate the size of the image to fit within a cell range
def resize_image_to_fit(ws, image_path, start_cell, end_cell):
    # Get the width and height of the target cell range
    col_start, row_start = ws[start_cell].column, ws[start_cell].row
    col_end, row_end = ws[end_cell].column, ws[end_cell].row

    # Calculate width and height in pixels
    width_in_pixels = sum([ws.column_dimensions[get_column_letter(col)].width or 8.43 for col in range(col_start, col_end + 1)]) * 7
    height_in_pixels = sum([ws.row_dimensions[row].height or 15 for row in range(row_start, row_end + 1)])

    # Open the image
    img = PilImage.open(image_path)

    # Convert the image to RGB if it has an alpha channel (RGBA)
    if img.mode == 'RGBA':
        img = img.convert('RGB')

    # Resize the image to fit within the calculated width and height
    img = img.resize((int(width_in_pixels), int(height_in_pixels)), PilImage.Resampling.LANCZOS)
    
    # Create the resized image path
    directory, filename = os.path.split(image_path)
    name, ext = os.path.splitext(filename)
    resized_image_path = os.path.join(directory, f"resized_{name}{ext}")
    
    # Save the resized image
    img.save(resized_image_path)

    return resized_image_path

# Function to insert order details in specific cells
def insert_order_details(ws, order_details):
    ws['C4'] = order_details.get('order_id')
    ws['C5'] = order_details.get('input_person')
    ws['E5'] = order_details.get('order_finish_time')
    ws['I5'] = order_details.get('inherit_id')
    ws['L5'] = order_details.get('customer_id')
    ws['O5'] = order_details.get('last_type')

# Function to insert series data starting from row 9
def insert_series_data(ws, series_data, start_row=9):
    for i, row_data in enumerate(series_data):
        for col_index, (col, value) in enumerate(row_data.items()):
            column_map = {
                "序号": "A", "材料类型": "B","材料名称": "C", "材料型号":"E", "材料规格": "G", "颜色": "I",
                "单位": "J", "厂家名称": "K", "单位用量": "L", "核定用量": "M", "使用工段": "N", "备注": "O"
            }
            column_letter = column_map.get(col)
            if column_letter:
                ws[f"{column_letter}{start_row + i}"] = value

# Function to insert an image into the Excel file
def insert_image(ws, image_path, start_cell, end_cell):
    # Resize the image to fit the target cell range
    resized_image_path = resize_image_to_fit(ws, image_path, start_cell, end_cell)
    
    # Insert the resized image into the worksheet
    img = Image(resized_image_path)
    ws.add_image(img, start_cell)

# Function to save the workbook after modification
def save_workbook(wb, new_file_path):
    wb.save(new_file_path)

# Main function to generate the Excel file
def generate_excel_file(template_path, new_file_path, order_details, series_data, image_url, image_save_path):
    # Load template
    wb, ws = load_template(template_path, new_file_path)

    # Insert the order details
    insert_order_details(ws, order_details)

    # Insert series data starting from row 9
    insert_series_data(ws, series_data, start_row=9)
    # Download the image from the URL
    download_image(image_url, image_save_path)
    if os.path.exists(image_save_path):
        insert_image(ws, image_save_path, "N1", "P3")

    # Save the modified workbook
    save_workbook(wb, new_file_path)

    print(f"Workbook saved as {new_file_path}")

# Example usage
# if __name__ == "__main__":
#     # Example parameters
#     template_path = "BOM-V1.0-temp.xlsx"  # Replace with your template path
#     new_file_path = "BOM-V1.0-modified.xlsx"  # The output file
#     image_url = "http://localhost:12667/shoe/3E3056/shoe_image.jpg"
#     image_save_path = "shoe_image.jpg"  # The image will be downloaded to this path

#     # Example order details
#     order_details = {
#         'order_id': '12345',
#         'input_person': 'John Doe',
#         'order_finish_time': '2024-09-01',
#         'inherit_id': '98765',
#         'customer_id': 'CUST123',
#         'last_type': 'TypeX'
#     }

#     # Example series data to be inserted starting from row 9
#     series_data = [
#         {"序号": "1", "材料类型": "Type1", "材料名称": "Material1", "材料规格": "Spec1", "颜色": "Red", "单位": "Unit1", "厂家名称": "Supplier1", "单位用量": "100", "核定用量": "90", "使用工段": "Step1", "备注": "None"},
#         {"序号": "2", "材料类型": "Type2", "材料名称": "Material2", "材料规格": "Spec2", "颜色": "Blue", "单位": "Unit2", "厂家名称": "Supplier2", "单位用量": "200", "核定用量": "180", "使用工段": "Step2", "备注": "Urgent"}
#     ]

#     # Call the function to generate the Excel file
#     generate_excel_file(template_path, new_file_path, order_details, series_data, image_url, image_save_path)
